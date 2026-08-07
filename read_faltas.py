import openpyxl

FILE = r"M:\ArchivosCompartidosResguardo\0_Jose Romero e hijos SRL\1_Produccion\2 - Laboratorio\Estadísticas\Indicadores de calidad\Faltas Personal.xlsm"

wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws = wb['DATOS']

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 8: break
    print(f"Fila {i}: {row}")

print(f"\nMax row: {ws.max_row}, Max col: {ws.max_column}")
wb.close()
